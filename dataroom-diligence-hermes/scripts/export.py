#!/usr/bin/env python3
"""Render scored findings as an XLSX issues list and a red flag memo.

Mechanical formatting only. Two safety habits, because contract text is
attacker-influenceable:

  * Spreadsheet cells that begin with = + - @ are prefixed with an apostrophe so
    Excel treats them as text rather than as a formula.
  * Extracted text is never used to build a path, a command, or a template
    directive. Template placeholders are replaced with rendered text, and the
    replacement runs once so injected placeholder syntax cannot re-expand.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

sys.path.insert(0, str(Path(__file__).resolve().parent))

from schemas import (  # noqa: E402
    CLAUSE_TYPES,
    SEVERITIES,
    SEVERITY_RANK,
    die,
    eprint,
    read_jsonl,
    require,
)

DEFAULT_TEMPLATE = (
    Path(__file__).resolve().parent.parent / "templates" / "redflag-memo.md"
)

COLUMNS = [
    ("doc_id", "Document", 26),
    ("doc_title", "Title", 34),
    ("page", "Page", 7),
    ("clause_type", "Clause type", 24),
    ("severity", "Severity", 18),
    ("severity_reason", "Why", 42),
    ("confidence", "Confidence", 11),
    ("low_confidence", "Low conf.", 10),
    ("quote", "Quote (verified)", 70),
    ("summary", "Summary", 50),
    ("terms_rendered", "Terms", 40),
    ("char_start", "Char start", 11),
    ("char_end", "Char end", 11),
    ("finding_id", "Finding id", 22),
]

_CONTROL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(value: Any) -> str:
    """Flatten a value to a single-line string safe to place in a cell."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    text = value if isinstance(value, str) else str(value)
    text = _CONTROL_CHARS_RE.sub(" ", text)
    return " ".join(text.split())


def spreadsheet_safe(value: Any) -> Any:
    """Neutralize formula injection from document text."""
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def render_terms(terms: Dict[str, Any]) -> str:
    if not isinstance(terms, dict) or not terms:
        return ""
    parts = []
    for key in sorted(terms):
        value = terms[key]
        if value is None or value == "":
            continue
        if isinstance(value, bool):
            rendered = "true" if value else "false"
        elif isinstance(value, (int, float)):
            rendered = str(value)
        else:
            rendered = clean(value)
        parts.append(f"{key}={rendered}")
    return "; ".join(parts)


def load_findings(path: Path) -> List[Dict[str, Any]]:
    findings = []
    for payload in read_jsonl(path):
        if not payload.get("validated"):
            die(
                f"{path} contains an unvalidated finding "
                f"({payload.get('finding_id')}). Run validate.py before export; "
                "unverified citations must never be reported."
            )
        payload["terms_rendered"] = render_terms(payload.get("terms") or {})
        findings.append(payload)
    findings.sort(
        key=lambda r: (
            SEVERITY_RANK.get(r.get("severity") or "note", 9),
            str(r.get("doc_id")),
            int(r.get("page") or 0),
            int(r.get("char_start") or 0),
        )
    )
    return findings


# --------------------------------------------------------------------------
# XLSX
# --------------------------------------------------------------------------


def write_xlsx(findings: List[Dict[str, Any]], path: Path) -> None:
    openpyxl = require("openpyxl", "openpyxl", "XLSX export")
    from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
    from openpyxl.utils import get_column_letter  # noqa: E402

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Issues"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F3864")
    severity_fill = {
        "blocking": PatternFill("solid", start_color="F4CCCC"),
        "consent_required": PatternFill("solid", start_color="FCE5CD"),
        "note": PatternFill("solid", start_color="EFEFEF"),
    }

    for index, (_key, label, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row_index, finding in enumerate(findings, start=2):
        severity = finding.get("severity") or "note"
        for column_index, (key, _label, _width) in enumerate(COLUMNS, start=1):
            raw = finding.get(key)
            if key in ("page", "char_start", "char_end"):
                value: Any = int(raw) if raw is not None else None
            elif key == "confidence":
                value = float(raw) if raw is not None else None
            else:
                value = spreadsheet_safe(clean(raw))
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=key in ("quote", "summary", "terms_rendered"))
            if column_index == 5 and severity in severity_fill:
                cell.fill = severity_fill[severity]

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNS))}{max(len(findings) + 1, 2)}"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


# --------------------------------------------------------------------------
# Memo
# --------------------------------------------------------------------------


def markdown_table(header: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    if not rows:
        return "_none_"
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("|" + "|".join(["---"] * len(header)) + "|")
    for row in rows:
        lines.append("| " + " | ".join(clean(cell) for cell in row) + " |")
    return "\n".join(lines)


def severity_table(findings: List[Dict[str, Any]]) -> str:
    counts = {name: 0 for name in SEVERITIES}
    for finding in findings:
        name = finding.get("severity") or "note"
        counts[name] = counts.get(name, 0) + 1
    rows = [[name, counts.get(name, 0)] for name in SEVERITIES]
    return markdown_table(["Severity", "Count"], rows)


def clause_table(findings: List[Dict[str, Any]]) -> str:
    rows = []
    for clause in CLAUSE_TYPES:
        subset = [f for f in findings if f.get("clause_type") == clause]
        if not subset:
            continue
        blocking = sum(1 for f in subset if f.get("severity") == "blocking")
        consent = sum(1 for f in subset if f.get("severity") == "consent_required")
        rows.append(
            [
                f"`{clause}`",
                len(subset),
                blocking,
                consent,
                len({f.get("doc_id") for f in subset}),
            ]
        )
    return markdown_table(
        ["Clause type", "Findings", "Blocking", "Consent required", "Documents"], rows
    )


def findings_sections(findings: List[Dict[str, Any]]) -> str:
    blocks: List[str] = []
    for clause in CLAUSE_TYPES:
        subset = [f for f in findings if f.get("clause_type") == clause]
        if not subset:
            continue
        blocks.append(f"### `{clause}` — {len(subset)} finding(s)\n")
        for finding in subset:
            title = clean(finding.get("doc_title")) or clean(finding.get("doc_id"))
            flag = " · low confidence" if finding.get("low_confidence") else ""
            confidence = finding.get("confidence")
            confidence_text = (
                f" · confidence {float(confidence):.2f}" if confidence is not None else ""
            )
            blocks.append(
                f"**{clean(finding.get('severity'))}** — {title} "
                f"(`{clean(finding.get('doc_id'))}`, p. {clean(finding.get('page'))}, "
                f"chars {clean(finding.get('char_start'))}–"
                f"{clean(finding.get('char_end'))}){confidence_text}{flag}"
            )
            blocks.append("")
            blocks.append(f"> {clean(finding.get('quote'))}")
            blocks.append("")
            blocks.append(f"{clean(finding.get('summary'))}")
            terms = clean(finding.get("terms_rendered"))
            if terms:
                blocks.append("")
                blocks.append(f"Terms: {terms}")
            reason = clean(finding.get("severity_reason"))
            if reason:
                blocks.append("")
                blocks.append(f"Severity rationale: {reason}")
            blocks.append("")
    return "\n".join(blocks).strip() or "_No findings._"


def validation_note(report_path: Optional[Path]) -> str:
    if not report_path:
        return (
            "All quotations were verified by `scripts/validate.py` against source "
            "text before export. Pass `--validation-report` to include the drop "
            "rate here."
        )
    if not report_path.exists():
        return f"_Validation report not found at {report_path}._"
    try:
        report = json.loads(report_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        return f"_Validation report unreadable: {exc}._"
    total = report.get("total", 0)
    validated = report.get("validated", 0)
    dropped = report.get("dropped", 0)
    rate = report.get("drop_rate", 0)
    reasons = report.get("by_reason") or {}
    lines = [
        f"`validate.py` checked {total} extracted findings against source text: "
        f"{validated} verified, {dropped} dropped ({float(rate) * 100:.1f}%).",
    ]
    if reasons:
        lines.append("")
        lines.append(
            markdown_table(
                ["Drop reason", "Count"],
                sorted(reasons.items(), key=lambda kv: -kv[1]),
            )
        )
    return "\n".join(lines)


def render_memo(
    template_text: str,
    findings: List[Dict[str, Any]],
    deal_structure: str,
    matter: str,
    generated_at: str,
    report_path: Optional[Path],
) -> str:
    replacements = {
        "{{MATTER}}": clean(matter),
        "{{DEAL_STRUCTURE}}": clean(deal_structure),
        "{{DOC_COUNT}}": str(len({f.get("doc_id") for f in findings})),
        "{{FINDING_COUNT}}": str(len(findings)),
        "{{GENERATED_AT}}": generated_at,
        "{{SEVERITY_TABLE}}": severity_table(findings),
        "{{CLAUSE_TABLE}}": clause_table(findings),
        "{{FINDINGS_BY_CLAUSE}}": findings_sections(findings),
        "{{VALIDATION_NOTE}}": validation_note(report_path),
    }
    # Single left-to-right pass over the template: rendered content is never
    # rescanned, so placeholder-looking text inside a contract quote is inert.
    pattern = re.compile("|".join(re.escape(key) for key in replacements))
    return pattern.sub(lambda match: replacements[match.group(0)], template_text)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export scored findings to XLSX and a red flag memo."
    )
    parser.add_argument("--input", required=True, help="scored.jsonl from score.py")
    parser.add_argument("--xlsx", required=True, help="issues.xlsx to write")
    parser.add_argument("--memo", required=True, help="redflag-memo.md to write")
    parser.add_argument(
        "--template",
        default=str(DEFAULT_TEMPLATE),
        help="memo template (default: ../templates/redflag-memo.md)",
    )
    parser.add_argument(
        "--matter", default="Target company data room", help="memo title"
    )
    parser.add_argument(
        "--deal-structure",
        default="",
        help="override the deal structure label (default: read from findings)",
    )
    parser.add_argument(
        "--validation-report",
        default="",
        help="validation-report.json from validate.py, summarized in the memo",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    findings = load_findings(Path(args.input))
    if not findings:
        die(f"{args.input} contained no findings; nothing to export", code=2)

    template_path = Path(args.template).expanduser()
    if not template_path.exists():
        die(f"memo template not found: {template_path}")

    structures = {f.get("deal_structure") for f in findings if f.get("deal_structure")}
    deal_structure = args.deal_structure or (
        structures.pop() if len(structures) == 1 else "unspecified"
    )

    write_xlsx(findings, Path(args.xlsx))

    memo = render_memo(
        template_path.read_text(encoding="utf-8"),
        findings,
        deal_structure,
        args.matter,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        Path(args.validation_report) if args.validation_report else None,
    )
    memo_path = Path(args.memo)
    memo_path.parent.mkdir(parents=True, exist_ok=True)
    memo_path.write_text(memo, encoding="utf-8")

    eprint(
        f"[export] {len(findings)} findings -> {args.xlsx} and {args.memo} "
        f"({deal_structure} deal)"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
